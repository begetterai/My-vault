#!/usr/bin/env python3
"""
Сравнивает Super P&L и Main P&L построчно за указанные месяцы.
Показывает только расхождения.
"""
import json, os, sys, io
from google.oauth2 import service_account
from google.auth.transport.requests import AuthorizedSession
import openpyxl

CREDS = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                     'credentials', 'romashka-drive.json')
SCOPES = ['https://www.googleapis.com/auth/spreadsheets',
          'https://www.googleapis.com/auth/drive.readonly']

MAIN_PNL_ID  = '10U6n12I9O9HJ4nRqXsmmyf2AOMb4Es9w'
SUPER_PNL_ID = '1l8Lau8K9997pyqJj-zjlILAkLoOHlQmPo6BAwqc5FBU'

SHEET_MAP = {
    'Лохути 11 (ЗБ)':          'ЗБ',
    'М. Турсунзода 12 (ОВИР)': 'ОВИР',
}

MONTHS_RU = ['','Январь','Февраль','Март','Апрель','Май','Июнь',
             'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']

# Строки для сравнения (данные, не формулы)
DATA_ROWS = [2, 4, 5, 7, 8, 13, 15, 18, 22, 23, 24,
             25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37,
             40, 44, 45, 46]

# Строки с формулами — тоже проверяем итоговые значения
FORMULA_ROWS = [3, 6, 9, 11, 17, 20, 38, 42, 47]

ALL_ROWS = sorted(DATA_ROWS + FORMULA_ROWS)


def col_letter(month):
    return chr(ord('B') + month)  # 1→C, ..., 12→N


def download_excel(session):
    r = session.get(
        f'https://www.googleapis.com/drive/v3/files/{MAIN_PNL_ID}',
        params={'alt': 'media', 'supportsAllDrives': 'true'},
        timeout=60
    )
    if r.status_code != 200:
        raise RuntimeError(f"Drive download: {r.status_code}")
    return io.BytesIO(r.content)


def read_super_pnl(session, sheet_name, months):
    cols = [col_letter(m) for m in months]
    ranges = [f"'{sheet_name}'!{c}1:{c}47" for c in cols]
    r = session.get(
        f'https://sheets.googleapis.com/v4/spreadsheets/{SUPER_PNL_ID}/values:batchGet',
        params={'ranges': ranges, 'valueRenderOption': 'UNFORMATTED_VALUE'},
        timeout=30
    )
    data = r.json().get('valueRanges', [])
    # result[row][month_idx]
    result = {}
    for mi, vr in enumerate(data):
        vals = vr.get('values', [])
        for ri, row_vals in enumerate(vals):
            row_num = ri + 1
            val = row_vals[0] if row_vals else None
            if row_num not in result:
                result[row_num] = {}
            result[row_num][months[mi]] = val
    return result


def num(v):
    if v is None or v == '':
        return None
    try:
        return float(str(v).replace(' ', '').replace('\xa0', '').replace(',', '.'))
    except (ValueError, TypeError):
        return None


def main(month_from=1, month_to=4):
    creds = service_account.Credentials.from_service_account_file(CREDS, scopes=SCOPES)
    session = AuthorizedSession(creds)

    print("Скачиваю Main P&L...")
    wb = openpyxl.load_workbook(download_excel(session), data_only=True)

    months = list(range(month_from, month_to + 1))

    for main_sheet, super_sheet in SHEET_MAP.items():
        print(f"\n{'═'*60}")
        print(f"  {main_sheet}  →  {super_sheet}")
        print(f"{'═'*60}")

        if main_sheet not in wb.sheetnames:
            print(f"  ⚠️  Лист не найден в Main P&L")
            continue

        ws = wb[main_sheet]
        super_data = read_super_pnl(session, super_sheet, months)

        # Читаем метки из Super P&L col B
        labels_r = session.get(
            f'https://sheets.googleapis.com/v4/spreadsheets/{SUPER_PNL_ID}/values/{super_sheet}!B1:B47',
            timeout=20
        )
        labels = [v[0] if v else '' for v in labels_r.json().get('values', [])]

        diffs = []
        matches = 0

        for row in ALL_ROWS:
            label = labels[row-1] if row-1 < len(labels) else f'R{row}'
            for m in months:
                col = col_letter(m)
                excel_col_idx = ord(col) - ord('A') + 1
                main_val = num(ws.cell(row=row, column=excel_col_idx).value)
                super_val = num(super_data.get(row, {}).get(m))

                if main_val is None and super_val is None:
                    continue
                if main_val is None:
                    main_val = 0.0
                if super_val is None:
                    super_val = 0.0

                diff = super_val - main_val
                if abs(diff) < 1:  # погрешность 1 сом
                    matches += 1
                else:
                    diffs.append((row, m, label, main_val, super_val, diff))

        if not diffs:
            print(f"  ✅ Всё совпадает ({matches} ячеек)")
        else:
            print(f"  ✅ Совпадает: {matches} ячеек")
            print(f"  ❌ Расхождений: {len(diffs)}")
            print()
            print(f"  {'Стр':>3}  {'Месяц':<10}  {'Метка':<30}  {'Main':>10}  {'Super':>10}  {'Разница':>10}")
            print(f"  {'-'*78}")
            for row, m, label, mv, sv, diff in sorted(diffs):
                sign = '+' if diff > 0 else ''
                print(f"  {row:>3}  {MONTHS_RU[m]:<10}  {label:<30}  {mv:>10,.0f}  {sv:>10,.0f}  {sign}{diff:>9,.0f}")


if __name__ == '__main__':
    args = sys.argv[1:]
    if not args:
        main()
    elif len(args) == 2:
        main(int(args[0]), int(args[1]))
    else:
        print("Использование: compare_pnl.py [месяц_от] [месяц_до]")
