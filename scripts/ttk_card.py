#!/usr/bin/env python3
"""Генератор ТТК-карточки А4 (полуфабрикат) в формате шаблона Ромашки.
Стиль зашит по эталону «Соус Спайси»: Arial, тёмно-синяя шапка, золотые секции, A4 альбомная."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# палитра шаблона
NAVY='FF1F2A37'; GOLD='FFB08D3E'; CREAM='FFF3EEE0'; INK='FF2B2B2B'; GRAY='FF8A8578'; WHITE='FFFFFFFF'
thin=Side(style='thin', color='FFBFAF8E')
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def _fill(c): return PatternFill('solid', fgColor=c)
def _f(sz=13,bold=False,italic=False,color=INK,name='Times New Roman'):
    # правило оформления Азиза: Times New Roman, размер 13 (жирный допустим)
    return Font(name=name,size=13,bold=bold,italic=italic,color=color)

def build_card(spec, path):
    wb=openpyxl.Workbook(); ws=wb.active
    ws.title=spec.get('sheet','TTK')
    ws.sheet_view.showGridLines=False
    # ширины колонок A..G (шаблон)
    for col,w in {'A':4,'B':26,'C':6,'D':8,'E':20,'F':24,'G':8}.items():
        ws.column_dimensions[col].width=w
    def cell(coord,val,font,fill=WHITE,align=('center','center'),wrap=True):
        c=ws[coord]; c.value=val; c.font=font; c.fill=_fill(fill)
        c.alignment=Alignment(horizontal=align[0],vertical=align[1],wrap_text=wrap); c.border=BORDER
        return c
    def bar_range(r0,r1,cols='ABCDEFGHIJKL',color=WHITE):
        for r in range(r0,r1+1):
            for col in cols:
                cc=ws[f'{col}{r}']; cc.fill=_fill(color); cc.border=BORDER
    ncols='ABCDEFGHIJKL'
    ings=spec['ingredients']; ni=len(ings)
    last_ing=3+ni  # строка последнего ингредиента
    # шапка
    ws.merge_cells(f'A1:F1'); ws.merge_cells(f'G1:L{last_ing}')
    ws.merge_cells('A2:F2')
    cell('A1',spec['title'],_f(18,bold=True,color=WHITE),NAVY); ws.row_dimensions[1].height=34
    bar_range(1,last_ing,'GHIJKL')
    cell('G1','МЕСТО ДЛЯ ФОТО ГОТОВОГО БЛЮДА',_f(10,italic=True,color=GRAY),WHITE)
    cell('A2',f"ВЫХОД: {spec['output']}",_f(11,bold=True),GOLD); ws.row_dimensions[2].height=22
    # заголовок таблицы
    heads=['№','Продукт / ПФ','Ед.','Вес','Нарезка / форма','Примечание']
    for col,h in zip('ABCDEF',heads):
        cell(f'{col}3',h,_f(9,bold=True),CREAM);
    ws.row_dimensions[3].height=18
    # ингредиенты
    for i,ing in enumerate(ings):
        r=4+i; ws.row_dimensions[r].height=20
        cell(f'A{r}',ing.get('n',i+1),_f(9))
        cell(f'B{r}',ing['name'],_f(9))
        cell(f'C{r}',ing.get('unit','г'),_f(9))
        cell(f'D{r}',ing['weight'],_f(9))
        cell(f'E{r}',ing.get('form','—'),_f(9))
        cell(f'F{r}',ing.get('note',''),_f(9))
    r=last_ing+1
    def section(title):
        nonlocal r
        ws.merge_cells(f'A{r}:L{r}')
        bar_range(r,r,color=GOLD)
        cell(f'A{r}',title,_f(10,bold=True,color=WHITE),GOLD,align=('left','center'))
        ws.row_dimensions[r].height=20; r+=1
    def line(text):
        nonlocal r
        ws.merge_cells(f'A{r}:L{r}')
        cell(f'A{r}',text,_f(9),WHITE,align=('left','center')); bar_range(r,r); r+=1
    section('ТЕХНОЛОГИЯ ПРИГОТОВЛЕНИЯ')
    for i,step in enumerate(spec['tech'],1): line(f'{i}. {step}')
    section('АЛЛЕРГЕНЫ')
    for t in (spec['allergens'] if isinstance(spec['allergens'],list) else [spec['allergens']]): line(t)
    section('ОРГАНОЛЕПТИКА')
    for t in spec['organoleptic']: line(t)
    section('ХРАНЕНИЕ И ИСПОЛЬЗОВАНИЕ')
    for t in spec['storage']: line(t)
    # печать A4 альбомная, в одну страницу
    ws.page_setup.orientation='landscape'; ws.page_setup.paperSize=9
    ws.sheet_properties.pageSetUpPr=openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=1
    ws.page_margins.left=ws.page_margins.right=ws.page_margins.top=ws.page_margins.bottom=0.3
    wb.save(path)
    return path
