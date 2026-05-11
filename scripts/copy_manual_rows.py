#!/usr/bin/env python3
"""
Копирует ФОТ (строки 13,15), Погашение долгов (44), Инвестиции (45),
Дивиденды (46) из Main P&L (Excel) в Super P&L за указанные месяцы.
"""
import json, os, sys, io
from google.oauth2 import service_account
from google.auth.transport.requests import AuthorizedSession
import openpyxl

CREDS = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                     'credentials', 'romashka-drive.json')
SCOPES = ['https://www.googleapis.com/auth/spreadsheets',
          'https://www.googleapis.com/auth/drive.readonly']

MAIN_PNL_ID  = '10U6n12I9O9HJ4nRqXsmmyf2AOMb4Es9w'
SUPER_PNL_ID = '1l8Lau8K9997pyqJj-zjlILAkLoOHlQmPo6BAwqc5FBU'

# Excel sheet name → Super P&L sheet name
SHEET_MAP = {
    'Лохути 11 (ЗБ)':          'ЗБ',
    'М. Турсунзода 12 (ОВИР)': 'ОВИР',
}

# Строки для ручного копирования (не из Poster)
# 44, 45, 46 — теперь берутся из Poster (finance.getTransactions)
MANUAL_ROWS = [13, 15]


def col_letter(month):
    """Месяц 1–12 → C–N"""
    return chr(ord('B') + month)


def download_excel(session):
    """Скачивает Excel файл из Drive."""
    r = session.get(
        f'https://www.googleapis.com/drive/v3/files/{MAIN_PNL_ID}',
        params={'alt': 'media', 'supportsAllDrives': 'true'},
        timeout=60
    )
    if r.status_code != 200:
        raise RuntimeError(f"Drive download failed: {r.status_code} {r.text[:300]}")
    return io.BytesIO(r.content)


def main(month_from=1, month_to=4):
    creds = service_account.Credentials.from_service_account_file(CREDS, scopes=SCOPES)
    session = AuthorizedSession(creds)

    print("Скачиваю Main P&L...")
    excel_bytes = download_excel(session)
    wb = openpyxl.load_workbook(excel_bytes, data_only=True)
    print(f"Листы в Main P&L: {wb.sheetnames}")

    months = list(range(month_from, month_to + 1))
    # В Main P&L: A=eng, B=ru, C=Январь, D=Февраль ... → col C = month 1
    # col index for month m: chr(ord('B') + m) → same as Super P&L

    for excel_name, super_name in SHEET_MAP.items():
        if excel_name not in wb.sheetnames:
            print(f"\n⚠️  Лист '{excel_name}' не найден")
            continue

        ws = wb[excel_name]
        print(f"\n{excel_name} → {super_name}")
        updates = []

        for row in MANUAL_ROWS:
            for m in months:
                col = col_letter(m)
                # In Excel: col C = column index 3
                excel_col_idx = ord(col) - ord('A') + 1
                cell_val = ws.cell(row=row, column=excel_col_idx).value
                if cell_val is None or cell_val == '':
                    continue
                try:
                    num = float(cell_val)
                except (TypeError, ValueError):
                    continue
                if num == 0:
                    continue
                updates.append({
                    'range': f"'{super_name}'!{col}{row}",
                    'values': [[round(num, 2)]]
                })
                print(f"  R{row} {col}: {num:,.0f}")

        if not updates:
            print(f"  ⚠️  Нет данных для копирования")
            continue

        body = {'valueInputOption': 'USER_ENTERED', 'data': updates}
        r2 = session.post(
            f'https://sheets.googleapis.com/v4/spreadsheets/{SUPER_PNL_ID}/values:batchUpdate',
            headers={'Content-Type': 'application/json'},
            data=json.dumps(body),
            timeout=30
        )
        ok = r2.status_code == 200
        print(f"  {'✅' if ok else '❌'} {len(updates)} ячеек записано")
        if not ok:
            print(f"  {r2.text[:300]}")


if __name__ == '__main__':
    args = sys.argv[1:]
    if len(args) == 0:
        main()
    elif len(args) == 2:
        main(int(args[0]), int(args[1]))
    else:
        print("Использование: copy_manual_rows.py [месяц_от] [месяц_до]")
