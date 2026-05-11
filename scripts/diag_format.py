#!/usr/bin/env python3
"""Диагностика: проверяет E4, C8, D22 в ЗБ и форматы из Main P&L."""
import json, os, io
from google.oauth2 import service_account
from google.auth.transport.requests import AuthorizedSession
import openpyxl

CREDS = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                     'credentials', 'romashka-drive.json')
SCOPES = ['https://www.googleapis.com/auth/spreadsheets',
          'https://www.googleapis.com/auth/drive.readonly']

MAIN_PNL_ID  = '10U6n12I9O9HJ4nRqXsmmyf2AOMb4Es9w'
SUPER_PNL_ID = '1l8Lau8K9997pyqJj-zjlILAkLoOHlQmPo6BAwqc5FBU'

def main():
    creds = service_account.Credentials.from_service_account_file(CREDS, scopes=SCOPES)
    session = AuthorizedSession(creds)

    # 1. Локаль Super P&L
    r = session.get(f'https://sheets.googleapis.com/v4/spreadsheets/{SUPER_PNL_ID}',
                    params={'fields': 'properties'}, timeout=20)
    props = r.json().get('properties', {})
    print(f"Super P&L locale: {props.get('locale')}, timeZone: {props.get('timeZone')}")

    # 2. Текущие значения проблемных ячеек в Super P&L
    ranges = ["'ЗБ'!E4", "'ЗБ'!C8", "'ЗБ'!D22"]
    r2 = session.get(
        f'https://sheets.googleapis.com/v4/spreadsheets/{SUPER_PNL_ID}/values:batchGet',
        params={'ranges': ranges, 'valueRenderOption': 'UNFORMATTED_VALUE'},
        timeout=20)
    print("\nSuper P&L проблемные ячейки:")
    for vr in r2.json().get('valueRanges', []):
        vals = vr.get('values', [[None]])
        print(f"  {vr['range']}: {vals[0][0] if vals else None}")

    # 3. Main P&L: те же ячейки
    r3 = session.get(
        f'https://www.googleapis.com/drive/v3/files/{MAIN_PNL_ID}',
        params={'alt': 'media', 'supportsAllDrives': 'true'}, timeout=60)
    wb = openpyxl.load_workbook(io.BytesIO(r3.content), data_only=True)
    ws = wb['Лохути 11 (ЗБ)']
    print("\nMain P&L те же ячейки:")
    # E4 = col E row 4; C8 = col C row 8; D22 = col D row 22
    for cell_addr in ['E4', 'C8', 'D22']:
        c = ws[cell_addr]
        print(f"  {cell_addr}: value={c.value}  format={c.number_format}")

    # 4. Форматы % строк в Main P&L (строки 10,12,14,16,19,21,39,41,43)
    print("\nMain P&L форматы % строк (col C):")
    pct_rows = [10, 12, 14, 16, 19, 21, 39, 41, 43]
    for row in pct_rows:
        c = ws.cell(row=row, column=3)
        print(f"  R{row} C: value={c.value}  format={c.number_format}")

    # 5. Форматы числовых строк в Main P&L (выручка, COGS)
    print("\nMain P&L форматы числовых строк (col C):")
    for row in [2, 3, 4, 9, 17, 38, 42]:
        c = ws.cell(row=row, column=3)
        print(f"  R{row} C: value={c.value}  format={c.number_format}")

if __name__ == '__main__':
    main()
